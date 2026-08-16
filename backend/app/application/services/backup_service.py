"""
Serviço de Backup e Exportação (D3).

Exporta os dados de uma empresa (multi-tenant) em três formatos:
  - Excel (.xlsx) — uma aba por módulo, cabeçalhos formatados
  - CSV — um arquivo .csv por módulo, empacotados num .zip
  - Backup JSON — um único .json com todas as tabelas (restaurável)

Tudo roda em memória (BytesIO) — nada é gravado em disco.
O filtro por empresa_id garante isolamento entre tenants.

Camada: Application.
"""
from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from uuid import UUID

from sqlalchemy import inspect as sa_inspect
from sqlalchemy.orm import Session

from app.core.config import get_settings

from app.infrastructure.database.models.atendimento import AtendimentoModel
from app.infrastructure.database.models.banco import ContaBancariaModel, LancamentoBancarioModel
from app.infrastructure.database.models.cliente import ClienteModel
from app.infrastructure.database.models.compra import CompraModel
from app.infrastructure.database.models.conta_pagar import ContaPagarModel
from app.infrastructure.database.models.conta_receber import ContaReceberModel
from app.infrastructure.database.models.diario_obra import RegistroDiarioModel
from app.infrastructure.database.models.documento import DocumentoModel
from app.infrastructure.database.models.estoque import ItemEstoqueModel
from app.infrastructure.database.models.fornecedor import FornecedorModel
from app.infrastructure.database.models.historico_preco import HistoricoPrecoEstoqueModel
from app.infrastructure.database.models.obra import ObraModel
from app.infrastructure.database.models.orcamento import OrcamentoModel
from app.infrastructure.database.models.recorrencia import RecorrenciaFinanceiraModel
from app.infrastructure.database.models.venda import VendaModel

# Mapa: chave do módulo -> (label amigável, model)
# A ordem aqui é a ordem que aparece no Excel / no zip.
MODULOS: dict[str, tuple[str, Any]] = {
    "clientes": ("Clientes", ClienteModel),
    "obras": ("Obras", ObraModel),
    "fornecedores": ("Fornecedores", FornecedorModel),
    "orcamentos": ("Orçamentos", OrcamentoModel),
    "vendas": ("Vendas", VendaModel),
    "compras": ("Compras", CompraModel),
    "estoque": ("Estoque", ItemEstoqueModel),
    "historico_preco": ("Histórico de Preços", HistoricoPrecoEstoqueModel),
    "contas_pagar": ("Contas a Pagar", ContaPagarModel),
    "contas_receber": ("Contas a Receber", ContaReceberModel),
    "recorrencias": ("Recorrências", RecorrenciaFinanceiraModel),
    "contas_bancarias": ("Contas Bancárias", ContaBancariaModel),
    "lancamentos_bancarios": ("Lançamentos Bancários", LancamentoBancarioModel),
    "atendimentos": ("Atendimentos", AtendimentoModel),
    "diario_obra": ("Diário de Obra", RegistroDiarioModel),
    "documentos": ("Documentos", DocumentoModel),
}


def modulos_disponiveis() -> list[dict[str, str]]:
    """Lista dos módulos exportáveis (para a página de Backup escolher)."""
    return [{"chave": k, "label": v[0]} for k, v in MODULOS.items()]


def _serializar_valor(v: Any) -> Any:
    """Converte tipos do banco em algo serializável (CSV/JSON/Excel)."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, UUID):
        return str(v)
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False, default=str)
    if isinstance(v, bool):
        return "Sim" if v else "Não"
    return v


def _colunas(model: Any) -> list[str]:
    """Nomes das colunas do model, na ordem definida."""
    return [c.key for c in sa_inspect(model).columns]


def _linhas(db: Session, model: Any, empresa_id: UUID) -> list[dict[str, Any]]:
    """Busca todas as linhas de um model para a empresa, como dicts."""
    cols = _colunas(model)
    # empresa_id existe em todos exceto se for a própria empresa (não exportada)
    query = db.query(model)
    if hasattr(model, "empresa_id"):
        query = query.filter(model.empresa_id == empresa_id)
    resultados = []
    for row in query.all():
        resultados.append({c: getattr(row, c) for c in cols})
    return resultados


# ─── EXCEL ────────────────────────────────────────────────────────────────────

def gerar_excel(db: Session, empresa_id: UUID, modulos: list[str]) -> bytes:
    """Gera um .xlsx com uma aba por módulo selecionado."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # remove aba default

    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="left", vertical="center")

    for chave in modulos:
        if chave not in MODULOS:
            continue
        label, model = MODULOS[chave]
        cols = _colunas(model)
        linhas = _linhas(db, model, empresa_id)

        # Nome da aba: máx 31 chars, sem caracteres proibidos
        aba_nome = label[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=aba_nome)

        # Cabeçalho
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Dados
        for ri, linha in enumerate(linhas, start=2):
            for ci, col in enumerate(cols, start=1):
                ws.cell(row=ri, column=ci, value=_serializar_valor(linha[col]))

        # Largura de colunas (auto simplificado)
        for ci, col in enumerate(cols, start=1):
            maxlen = len(col)
            for linha in linhas[:100]:  # amostra
                val = str(_serializar_valor(linha[col]))
                maxlen = max(maxlen, len(val))
            ws.column_dimensions[get_column_letter(ci)].width = min(maxlen + 3, 50)

        ws.freeze_panes = "A2"  # trava cabeçalho

    # Se nenhuma aba foi criada, cria uma vazia (Excel exige >=1 aba)
    if not wb.sheetnames:
        wb.create_sheet(title="Vazio")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── CSV (zip) ────────────────────────────────────────────────────────────────

def gerar_csv_zip(db: Session, empresa_id: UUID, modulos: list[str]) -> bytes:
    """Gera um .zip contendo um .csv por módulo."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for chave in modulos:
            if chave not in MODULOS:
                continue
            label, model = MODULOS[chave]
            cols = _colunas(model)
            linhas = _linhas(db, model, empresa_id)

            csv_buf = io.StringIO()
            writer = csv.writer(csv_buf, delimiter=";")  # ; para Excel BR
            writer.writerow(cols)
            for linha in linhas:
                writer.writerow([_serializar_valor(linha[c]) for c in cols])

            # BOM para o Excel abrir acentos corretamente
            conteudo = "\ufeff" + csv_buf.getvalue()
            zf.writestr(f"{chave}.csv", conteudo.encode("utf-8"))

    return buf.getvalue()


# ─── BACKUP JSON (restaurável) ────────────────────────────────────────────────

def gerar_backup_json(db: Session, empresa_id: UUID) -> bytes:
    """
    Gera um backup completo em JSON, com TODAS as tabelas da empresa.
    Formato restaurável: { "versao": 1, "empresa_id": ..., "gerado_em": ...,
    "dados": { "<tabela>": [ {linha}, ... ] } }
    """
    dados: dict[str, list[dict[str, Any]]] = {}
    for chave, (_label, model) in MODULOS.items():
        cols = _colunas(model)
        linhas = _linhas(db, model, empresa_id)
        dados[chave] = [
            {c: _serializar_valor(linha[c]) for c in cols}
            for linha in linhas
        ]

    backup = {
        "versao": 1,
        "empresa_id": str(empresa_id),
        "gerado_em": datetime.utcnow().isoformat(),
        "total_registros": sum(len(v) for v in dados.values()),
        "dados": dados,
    }
    return json.dumps(backup, ensure_ascii=False, indent=2, default=str).encode("utf-8")


# ─── ARQUIVOS DO STORAGE (documentos) ─────────────────────────────────────────

def _extrair_caminho_storage(arquivo_url: str) -> str | None:
    """
    Extrai o caminho relativo dentro do bucket a partir da URL assinada.
    Formato: .../object/sign/documentos/<caminho>?token=...
    """
    import re
    from urllib.parse import unquote
    m = re.search(r"/object/sign/documentos/([^?]+)", arquivo_url)
    if not m:
        return None
    return unquote(m.group(1))


def _baixar_arquivo_storage(caminho: str) -> bytes | None:
    """
    Baixa um arquivo do bucket privado 'documentos' usando a service role key.
    Retorna os bytes ou None em caso de falha (best-effort).
    """
    import httpx
    settings = get_settings()
    base = settings.SUPABASE_URL.rstrip("/")
    url = f"{base}/storage/v1/object/documentos/{caminho}"
    headers = {
        "Authorization": f"Bearer {settings.SUPABASE_SERVICE_ROLE_KEY}",
        "apikey": settings.SUPABASE_SERVICE_ROLE_KEY,
    }
    try:
        with httpx.Client(timeout=30) as client:
            resp = client.get(url, headers=headers)
            if resp.status_code == 200:
                return resp.content
    except Exception:
        pass
    return None


def gerar_backup_com_arquivos(
    db: Session, empresa_id: UUID, modulos: list[str], formato_dados: str
) -> bytes:
    """
    Gera um .zip contendo:
      - os dados (em Excel OU CSVs, conforme formato_dados)
      - uma pasta 'documentos/' com os arquivos binários baixados do Storage

    formato_dados: 'excel' | 'csv'
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # 1. Dados
        if formato_dados == "csv":
            # Reaproveita o gerador de CSV-zip: extrai seus membros para cá
            csv_zip_bytes = gerar_csv_zip(db, empresa_id, modulos)
            with zipfile.ZipFile(io.BytesIO(csv_zip_bytes)) as inner:
                for name in inner.namelist():
                    zf.writestr(f"dados/{name}", inner.read(name))
        else:
            excel_bytes = gerar_excel(db, empresa_id, modulos)
            zf.writestr("dados/inovak-dados.xlsx", excel_bytes)

        # 2. Documentos do Storage
        docs = _linhas(db, DocumentoModel, empresa_id)
        usados: set[str] = set()
        manifesto = []
        for doc in docs:
            url = doc.get("arquivo_url")
            nome_original = doc.get("arquivo_nome") or "arquivo"
            if not url:
                continue
            caminho = _extrair_caminho_storage(str(url))
            if not caminho:
                continue
            conteudo = _baixar_arquivo_storage(caminho)
            if conteudo is None:
                manifesto.append(f"FALHOU: {nome_original} ({caminho})")
                continue
            # Evitar colisão de nomes no zip
            nome_zip = nome_original
            i = 1
            while nome_zip in usados:
                base_n, _, ext = nome_original.rpartition(".")
                nome_zip = f"{base_n}-{i}.{ext}" if ext else f"{nome_original}-{i}"
                i += 1
            usados.add(nome_zip)
            zf.writestr(f"documentos/{nome_zip}", conteudo)
            manifesto.append(f"OK: {nome_zip}")

        # 3. Manifesto (o que entrou no backup)
        cab = f"Backup Inovak com arquivos\nGerado em: {datetime.utcnow().isoformat()}\nTotal de documentos: {len(docs)}\n\n"
        zf.writestr("documentos/_MANIFESTO.txt", (cab + "\n".join(manifesto)).encode("utf-8"))

    return buf.getvalue()
